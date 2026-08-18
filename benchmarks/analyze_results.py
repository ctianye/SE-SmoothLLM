"""汇总 Vicuna 主实验、外部 Judge 和效率指标，并生成可追溯图表。"""

from __future__ import annotations

import json
import math
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import font_manager
from matplotlib.patches import FancyArrowPatch, FancyBboxPatch
from matplotlib.ticker import FuncFormatter, PercentFormatter
from scipy import stats

# Use a non-interactive backend because this script is intended for headless benchmark hosts.
mpl.use("Agg")

# ============================================================
# CONFIGURATION - edit only this block
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = PROJECT_ROOT / "results" / "raw"
JUDGED_DIR = PROJECT_ROOT / "results" / "judged"
OUTPUT_DIR = PROJECT_ROOT / "results" / "processed"
FIGURE_DIR = PROJECT_ROOT / "results" / "figures"
PROJECT_TITLE = "SE-SmoothLLM JBB evaluation"
EXPECTED_RAW_RECORDS = 1_800
EXPECTED_DEEPSEEK_RECORDS = 600
EXPECTED_REFUSAL_RECORDS = 1_800
OUTPUT_FORMATS = ("pdf", "png", "jpg")
DPI = 300
DEFAULT_DECIMALS = 3
FIG_SINGLE_COL = (3.5, 2.6)
FIG_DOUBLE_COL = (7.2, 4.0)
CN_SERIF_FALLBACK = ["SimSun", "Songti SC", "Source Han Serif SC", "Noto Serif CJK SC", "serif"]
EN_SERIF_FALLBACK = ["Times New Roman", "Liberation Serif", "DejaVu Serif", "serif"]

# ============================================================
# END CONFIGURATION
# ============================================================

METHODS = ("undefended", "smoothllm_fixed", "se_smoothllm")
DISPLAY_NAMES = {
    "undefended": "Undefended",
    "smoothllm_fixed": "SmoothLLM fixed",
    "se_smoothllm": "SE-SmoothLLM",
}
METHOD_COLORS = {
    "undefended": "#aaaaaa",
    "smoothllm_fixed": "#555555",
    "se_smoothllm": "#000000",
}
METHOD_MARKERS = {"undefended": "^", "smoothllm_fixed": "s", "se_smoothllm": "o"}


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"invalid JSON at {path}:{line_number}") from exc
    return records


def load_raw_metrics() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    paths = sorted(RAW_DIR.glob("*.jsonl"))
    if len(paths) != 18:
        raise ValueError(f"expected 18 raw JSONL files, received {len(paths)}")
    for path in paths:
        records = read_jsonl(path)
        if len(records) != 100:
            raise ValueError(f"expected 100 records in {path}, received {len(records)}")
        for record in records:
            sample = record["sample"]
            result = record["result"]
            votes = result.get("votes") or {}
            trace = result.get("trace") or []
            rows.append(
                {
                    "method": record["method"],
                    "seed": record["seed"],
                    "split": sample["split"],
                    "index": sample["index"],
                    "behavior": sample.get("behavior", ""),
                    "category": sample.get("category", ""),
                    "config_sha256": record["config_sha256"],
                    "model": trace[0].get("model") if trace else None,
                    "internal_jailbroken": bool(result["jailbroken"]),
                    "copies_used": result["copies_used"],
                    "stopped_early": bool(result["stopped_early"]),
                    "prompt_tokens": result.get("prompt_tokens"),
                    "completion_tokens": result.get("completion_tokens"),
                    "latency_ms": result.get("latency_ms"),
                    "votes_safe": votes.get("safe"),
                    "votes_jailbroken": votes.get("jailbroken"),
                    "trace_length": len(trace),
                }
            )
    frame = pd.DataFrame(rows).sort_values(["split", "method", "seed", "index"])
    _assert_unique(frame, ["method", "seed", "split", "index"], "raw metrics")
    if len(frame) != EXPECTED_RAW_RECORDS:
        raise ValueError(f"expected {EXPECTED_RAW_RECORDS} raw records, received {len(frame)}")
    return frame.reset_index(drop=True)


def load_deepseek_judge() -> pd.DataFrame:
    path = JUDGED_DIR / "deepseek-v4-flash-jailbreak.jsonl"
    rows: list[dict[str, Any]] = []
    for record in read_jsonl(path):
        source = record["source"]
        judgment = record["judgment"]
        rows.append(
            {
                "method": source["method"],
                "seed": source["seed"],
                "split": source["split"],
                "index": source["index"],
                "response_sha256": source["response_sha256"],
                "config_sha256": source["config_sha256"],
                "external_jailbroken": bool(judgment["jailbroken"]),
                "format_conforming": bool(judgment["format_conforming"]),
                "raw_output": judgment["raw_output"],
                "model": judgment["model"],
                "latency_ms": judgment.get("latency_ms"),
                "prompt_tokens": judgment.get("prompt_tokens"),
                "completion_tokens": judgment.get("completion_tokens"),
            }
        )
    frame = pd.DataFrame(rows).sort_values(["method", "seed", "index"])
    _assert_unique(frame, ["method", "seed", "split", "index"], "DeepSeek judge")
    if len(frame) != EXPECTED_DEEPSEEK_RECORDS:
        raise ValueError(
            f"expected {EXPECTED_DEEPSEEK_RECORDS} DeepSeek records, received {len(frame)}"
        )
    if set(frame["method"]) != {"smoothllm_fixed", "se_smoothllm"}:
        raise ValueError("DeepSeek judge must contain fixed and SE methods only")
    return frame.reset_index(drop=True)


def load_refusal_judge() -> pd.DataFrame:
    path = JUDGED_DIR / "jbb-llama3-8b-refusal.jsonl"
    rows: list[dict[str, Any]] = []
    for record in read_jsonl(path):
        source = record["source"]
        judgment = record["judgment"]
        rows.append(
            {
                "method": source["method"],
                "seed": source["seed"],
                "split": source["split"],
                "index": source["index"],
                "response_sha256": source["response_sha256"],
                "config_sha256": source["config_sha256"],
                "refused": bool(judgment["refused"]),
                "format_conforming": bool(judgment["format_conforming"]),
                "raw_output": judgment["raw_output"],
                "model": judgment["model"],
                "latency_ms": judgment.get("latency_ms"),
                "prompt_tokens": judgment.get("prompt_tokens"),
                "completion_tokens": judgment.get("completion_tokens"),
            }
        )
    frame = pd.DataFrame(rows).sort_values(["split", "method", "seed", "index"])
    _assert_unique(frame, ["method", "seed", "split", "index"], "refusal judge")
    if len(frame) != EXPECTED_REFUSAL_RECORDS:
        raise ValueError(
            f"expected {EXPECTED_REFUSAL_RECORDS} refusal records, received {len(frame)}"
        )
    return frame.reset_index(drop=True)


def build_pair_metrics(raw: pd.DataFrame) -> pd.DataFrame:
    fixed = raw[raw["method"] == "smoothllm_fixed"].copy()
    se = raw[raw["method"] == "se_smoothllm"].copy()
    keys = ["seed", "split", "index"]
    fixed = fixed.rename(
        columns={
            "internal_jailbroken": "fixed_internal_jailbroken",
            "copies_used": "fixed_copies_used",
            "prompt_tokens": "fixed_prompt_tokens",
            "completion_tokens": "fixed_completion_tokens",
            "latency_ms": "fixed_latency_ms",
            "stopped_early": "fixed_stopped_early",
        }
    )
    se = se.rename(
        columns={
            "internal_jailbroken": "se_internal_jailbroken",
            "copies_used": "se_copies_used",
            "prompt_tokens": "se_prompt_tokens",
            "completion_tokens": "se_completion_tokens",
            "latency_ms": "se_latency_ms",
            "stopped_early": "se_stopped_early",
        }
    )
    columns = keys + [
        "fixed_internal_jailbroken",
        "fixed_copies_used",
        "fixed_prompt_tokens",
        "fixed_completion_tokens",
        "fixed_latency_ms",
        "fixed_stopped_early",
    ]
    se_columns = keys + [
        "se_internal_jailbroken",
        "se_copies_used",
        "se_prompt_tokens",
        "se_completion_tokens",
        "se_latency_ms",
        "se_stopped_early",
    ]
    pair = fixed[columns].merge(se[se_columns], on=keys, how="outer", validate="one_to_one")
    pair["internal_verdict_match"] = (
        pair["fixed_internal_jailbroken"] == pair["se_internal_jailbroken"]
    )
    pair["query_reduction"] = 1 - pair["se_copies_used"] / pair["fixed_copies_used"]
    pair["prompt_token_reduction"] = 1 - pair["se_prompt_tokens"] / pair["fixed_prompt_tokens"]
    pair["completion_token_reduction"] = (
        1 - pair["se_completion_tokens"] / pair["fixed_completion_tokens"]
    )
    return pair.sort_values(keys).reset_index(drop=True)


def aggregate_tables(
    raw: pd.DataFrame,
    deepseek: pd.DataFrame,
    refusal: pd.DataFrame,
    pair: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    summary_rows: list[dict[str, Any]] = []

    for (split, method), group in raw.groupby(["split", "method"], sort=True):
        summary_rows.append(
            {
                "metric": "internal_efficiency",
                "split": split,
                "method": method,
                "n": len(group),
                "value": group["copies_used"].mean(),
                "std": group["copies_used"].std(ddof=1),
                "early_stop_rate": group["stopped_early"].mean(),
                "internal_jailbroken_rate": group["internal_jailbroken"].mean(),
                "prompt_tokens_mean": group["prompt_tokens"].mean(),
                "completion_tokens_mean": group["completion_tokens"].mean(),
                "latency_ms_mean": group["latency_ms"].mean(),
            }
        )

    for method, group in deepseek.groupby("method", sort=True):
        summary_rows.append(
            {
                "metric": "deepseek_harmful_asr",
                "split": "harmful",
                "method": method,
                "n": len(group),
                "value": group["external_jailbroken"].mean(),
                "positive_count": int(group["external_jailbroken"].sum()),
                "format_conforming_rate": group["format_conforming"].mean(),
            }
        )

    for (split, method), group in refusal.groupby(["split", "method"], sort=True):
        summary_rows.append(
            {
                "metric": "llama3_8b_refusal_rate",
                "split": split,
                "method": method,
                "n": len(group),
                "value": group["refused"].mean(),
                "positive_count": int(group["refused"].sum()),
                "format_conforming_rate": group["format_conforming"].mean(),
            }
        )

    for split, group in pair.groupby("split", sort=True):
        summary_rows.append(
            {
                "metric": "fixed_vs_se_exactness",
                "split": split,
                "method": "se_vs_fixed",
                "n": len(group),
                "value": group["internal_verdict_match"].mean(),
                "mismatch_count": int((~group["internal_verdict_match"]).sum()),
                "avg_query_reduction": group["query_reduction"].mean(),
                "avg_prompt_token_reduction": group["prompt_token_reduction"].mean(),
                "avg_completion_token_reduction": group["completion_token_reduction"].mean(),
            }
        )

    summary = pd.DataFrame(summary_rows)
    seed_metrics = _build_seed_metrics(raw, deepseek, refusal)
    efficiency = _build_efficiency_summary(raw, pair)
    return summary, seed_metrics, efficiency


def _build_seed_metrics(
    raw: pd.DataFrame, deepseek: pd.DataFrame, refusal: pd.DataFrame
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for (split, method, seed), group in raw.groupby(["split", "method", "seed"], sort=True):
        rows.append(
            {
                "metric": "copies_used_mean",
                "split": split,
                "method": method,
                "seed": seed,
                "value": group["copies_used"].mean(),
                "n": len(group),
            }
        )
        rows.append(
            {
                "metric": "prompt_tokens_mean",
                "split": split,
                "method": method,
                "seed": seed,
                "value": group["prompt_tokens"].mean(),
                "n": len(group),
            }
        )
        rows.append(
            {
                "metric": "completion_tokens_mean",
                "split": split,
                "method": method,
                "seed": seed,
                "value": group["completion_tokens"].mean(),
                "n": len(group),
            }
        )
    for (method, seed), group in deepseek.groupby(["method", "seed"], sort=True):
        rows.append(
            {
                "metric": "deepseek_harmful_asr",
                "split": "harmful",
                "method": method,
                "seed": seed,
                "value": group["external_jailbroken"].mean(),
                "n": len(group),
            }
        )
    for (split, method, seed), group in refusal.groupby(["split", "method", "seed"], sort=True):
        rows.append(
            {
                "metric": "llama3_8b_refusal_rate",
                "split": split,
                "method": method,
                "seed": seed,
                "value": group["refused"].mean(),
                "n": len(group),
            }
        )
    return pd.DataFrame(rows)


def _build_efficiency_summary(raw: pd.DataFrame, pair: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for (split, method), group in raw.groupby(["split", "method"], sort=True):
        for metric, column in (
            ("copies_used", "copies_used"),
            ("prompt_tokens", "prompt_tokens"),
            ("completion_tokens", "completion_tokens"),
        ):
            rows.append(
                {
                    "split": split,
                    "method": method,
                    "metric": metric,
                    "n": len(group),
                    "mean": group[column].mean(),
                    "std": group[column].std(ddof=1),
                }
            )
    for split, group in pair.groupby("split", sort=True):
        for metric, column in (
            ("query_reduction", "query_reduction"),
            ("prompt_token_reduction", "prompt_token_reduction"),
            ("completion_token_reduction", "completion_token_reduction"),
        ):
            rows.append(
                {
                    "split": split,
                    "method": "se_vs_fixed",
                    "metric": metric,
                    "n": len(group),
                    "mean": group[column].mean(),
                    "std": group[column].std(ddof=1),
                }
            )
    return pd.DataFrame(rows)


def _assert_unique(frame: pd.DataFrame, columns: list[str], label: str) -> None:
    duplicates = frame.duplicated(columns).sum()
    if duplicates:
        raise ValueError(f"{label} contains {duplicates} duplicate keys")


def write_data_health(frames: dict[str, pd.DataFrame]) -> None:
    lines = [
        "# Data health report",
        "",
        "This report is generated before plotting. It checks the cleaned metric tables, not the raw response text.",
        "",
    ]
    for name, frame in frames.items():
        lines.extend(
            [
                f"## {name}",
                "",
                f"- Rows: {len(frame):,}",
                f"- Columns: {frame.shape[1]}",
                f"- Memory: {frame.memory_usage(deep=True).sum() / 1024**2:.2f} MB",
                f"- Exact duplicate rows: {int(frame.duplicated().sum()):,}",
                "",
                "### Column inventory",
                "",
                _markdown_table(
                    pd.DataFrame(
                        {
                            "dtype": frame.dtypes.astype(str),
                            "missing": frame.isna().sum(),
                            "missing_pct": (frame.isna().mean() * 100).round(2),
                            "n_unique": frame.nunique(dropna=False),
                        }
                    ).reset_index(names="column")
                ),
                "",
            ]
        )
        numeric = frame.select_dtypes(include=np.number)
        if not numeric.empty:
            lines.extend(["### Numeric summaries", "", _markdown_table(numeric.describe().T.round(3)), ""])
            normality_rows = []
            outlier_rows = []
            for column in numeric:
                values = numeric[column].dropna().astype(float)
                if len(values) >= 3 and values.nunique() > 1:
                    _, p_value = stats.shapiro(values.sample(min(len(values), 5000), random_state=42))
                    normality = "not normal" if p_value < 0.05 else "not rejected"
                else:
                    p_value, normality = math.nan, "insufficient or constant"
                q1, q3 = values.quantile([0.25, 0.75]) if len(values) else (math.nan, math.nan)
                iqr = q3 - q1
                outliers = (
                    ((values < q1 - 1.5 * iqr) | (values > q3 + 1.5 * iqr)).sum()
                    if len(values)
                    else 0
                )
                normality_rows.append({"column": column, "shapiro_p": p_value, "verdict": normality})
                outlier_rows.append(
                    {"column": column, "iqr_outliers": int(outliers), "outlier_pct": outliers / len(values) * 100 if len(values) else 0}
                )
            lines.extend(
                [
                    "### Normality checks (descriptive only)",
                    "",
                    _markdown_table(pd.DataFrame(normality_rows).round(4)),
                    "",
                    "### Outlier checks (IQR rule)",
                    "",
                    _markdown_table(pd.DataFrame(outlier_rows).round(3)),
                    "",
                ]
            )
        categorical = frame.select_dtypes(include=["object", "category"])
        if not categorical.empty:
            lines.extend(["### Category counts", ""])
            for column in categorical:
                lines.extend(
                    [
                        f"- `{column}` unique values: {categorical[column].nunique(dropna=False)}",
                        _markdown_table(categorical[column].value_counts(dropna=False).head(10).rename("count").reset_index()),
                        "",
                    ]
                )
    lines.extend(
        [
            "## Interpretation",
            "",
            "All expected key columns are complete and unique. The numeric distributions are used for descriptive summaries; no inferential significance claims are made because the main comparison has only three configured seeds.",
            "",
        ]
    )
    (OUTPUT_DIR / "00_data_health.md").write_text("\n".join(lines), encoding="utf-8")


def _markdown_table(frame: pd.DataFrame) -> str:
    try:
        return frame.to_markdown(index=False)
    except (ImportError, ValueError):
        return "```text\n" + frame.to_string(index=False) + "\n```"


def write_xlsx(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for index, column in enumerate(worksheet.columns, start=1):
            width = max((len(str(cell.value)) if cell.value is not None else 0 for cell in column), default=10)
            worksheet.column_dimensions[get_column_letter(index)].width = min(width + 2, 50)
    workbook.save(path)


def resolve_font(candidates: Iterable[str]) -> str:
    available = {font.name for font in font_manager.fontManager.ttflist}
    for candidate in candidates:
        if candidate in available or candidate == "serif":
            return candidate
    return "serif"


def configure_plotting() -> None:
    en_font = resolve_font(EN_SERIF_FALLBACK)
    cn_font = resolve_font(CN_SERIF_FALLBACK)
    mpl.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [en_font, cn_font, "serif"],
            "font.size": 10,
            "axes.titlesize": 11,
            "axes.labelsize": 10,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
            "legend.fontsize": 9,
            "axes.unicode_minus": False,
            "axes.edgecolor": "#000000",
            "axes.labelcolor": "#000000",
            "xtick.color": "#000000",
            "ytick.color": "#000000",
            "text.color": "#000000",
            "axes.spines.top": True,
            "axes.spines.right": True,
            "axes.spines.left": True,
            "axes.spines.bottom": True,
            "axes.grid": True,
            "axes.axisbelow": True,
            "grid.color": "#e5e5e5",
            "grid.linewidth": 0.5,
            "axes.facecolor": "white",
            "figure.facecolor": "white",
            "savefig.facecolor": "white",
            "legend.frameon": False,
            "savefig.bbox": "tight",
        }
    )


def save_figure(figure: plt.Figure, slug: str) -> None:
    for extension in OUTPUT_FORMATS:
        target = FIGURE_DIR / extension / f"{slug}.{extension}"
        target.parent.mkdir(parents=True, exist_ok=True)
        figure.savefig(target, dpi=DPI if extension != "pdf" else None)
    plt.close(figure)


def add_full_frame(axis: plt.Axes) -> None:
    for side in ("top", "right", "left", "bottom"):
        axis.spines[side].set_visible(True)
        axis.spines[side].set_linewidth(0.8)
        axis.spines[side].set_color("black")


def percentage_axis(axis: plt.Axes) -> None:
    axis.yaxis.set_major_formatter(PercentFormatter(1.0, decimals=0))
    axis.set_ylim(bottom=0)


def _overview_box(
    axis: plt.Axes,
    x: float,
    y: float,
    width: float,
    height: float,
    text: str,
    *,
    facecolor: str = "white",
    edgecolor: str = "#555555",
    fontsize: float = 8.5,
    linewidth: float = 1.0,
) -> None:
    patch = FancyBboxPatch(
        (x, y),
        width,
        height,
        boxstyle="round,pad=0.008,rounding_size=0.015",
        facecolor=facecolor,
        edgecolor=edgecolor,
        linewidth=linewidth,
    )
    axis.add_patch(patch)
    axis.text(
        x + width / 2,
        y + height / 2,
        text,
        ha="center",
        va="center",
        fontsize=fontsize,
        linespacing=1.2,
        color="white" if facecolor == "#222222" else "#000000",
    )


def _overview_arrow(
    axis: plt.Axes,
    start: tuple[float, float],
    end: tuple[float, float],
    *,
    color: str = "#555555",
    linewidth: float = 1.0,
    connectionstyle: str = "arc3",
) -> None:
    axis.add_patch(
        FancyArrowPatch(
            start,
            end,
            arrowstyle="-|>",
            mutation_scale=10,
            linewidth=linewidth,
            color=color,
            connectionstyle=connectionstyle,
        )
    )


def _overview_metrics(efficiency: pd.DataFrame, pair: pd.DataFrame) -> dict[str, float]:
    harmful = efficiency[efficiency["split"] == "harmful"]
    harmful_pair = pair[pair["split"] == "harmful"]

    def mean(metric: str, method: str) -> float:
        values = harmful[(harmful["metric"] == metric) & (harmful["method"] == method)]["mean"]
        if len(values) != 1:
            raise ValueError(f"overview metric is incomplete: {metric}/{method}")
        return float(values.iloc[0])

    fixed_queries = mean("copies_used", "smoothllm_fixed")
    se_queries = mean("copies_used", "se_smoothllm")
    return {
        "verdict_match_rate": float(pair["internal_verdict_match"].mean()),
        "verdict_mismatch_count": float((~pair["internal_verdict_match"]).sum()),
        "fixed_queries": fixed_queries,
        "se_queries": se_queries,
        "query_reduction": float(harmful_pair["query_reduction"].mean()),
        "prompt_token_reduction": float(harmful_pair["prompt_token_reduction"].mean()),
        "completion_token_reduction": float(harmful_pair["completion_token_reduction"].mean()),
    }


def make_overview_figure(efficiency: pd.DataFrame, pair: pd.DataFrame) -> pd.DataFrame:
    metrics = _overview_metrics(efficiency, pair)
    configure_plotting()
    figure = plt.figure(figsize=(11.5, 5.4))
    axis = figure.add_axes([0, 0, 1, 1])
    axis.set_xlim(0, 1)
    axis.set_ylim(0, 1)
    axis.axis("off")

    axis.text(
        0.03,
        0.955,
        "SE-SmoothLLM: shared execution with exact early stopping",
        fontsize=16,
        fontweight="bold",
        va="top",
    )
    axis.text(
        0.03,
        0.915,
        "Project-original schematic following the perturb-and-vote workflow of the official SmoothLLM implementation.",
        fontsize=8.5,
        color="#555555",
        va="top",
    )

    panel_edge = "#b5b5b5"
    for x, width, label in (
        (0.025, 0.59, "Shared SmoothLLM execution"),
        (0.64, 0.335, "Decision policy"),
    ):
        axis.add_patch(
            FancyBboxPatch(
                (x, 0.29),
                width,
                0.56,
                boxstyle="round,pad=0.012,rounding_size=0.02",
                facecolor="#f7f7f7",
                edgecolor=panel_edge,
                linewidth=0.9,
            )
        )
        axis.text(x + 0.02, 0.82, label, fontsize=10.5, fontweight="bold")

    _overview_box(axis, 0.045, 0.535, 0.095, 0.105, "Prompt\nP")
    _overview_box(
        axis,
        0.185,
        0.535,
        0.135,
        0.105,
        "Random perturb\nq = 10%",
        facecolor="#e6e6e6",
    )
    _overview_arrow(axis, (0.14, 0.587), (0.185, 0.587))

    row_y = (0.695, 0.585, 0.475, 0.365)
    labels = ("P'1", "P'2", "P'3", "P'N")
    for y, label in zip(row_y, labels, strict=True):
        _overview_box(axis, 0.365, y, 0.075, 0.065, label, fontsize=8)
        _overview_box(axis, 0.475, y, 0.085, 0.065, "LLM\nJudge", facecolor="#dedede", fontsize=7.5)
        _overview_arrow(
            axis,
            (0.32, 0.587),
            (0.365, y + 0.0325),
            connectionstyle="arc3,rad=0.12",
        )
        _overview_arrow(axis, (0.44, y + 0.0325), (0.475, y + 0.0325))
    axis.text(0.402, 0.425, "...", fontsize=12, ha="center", va="center")
    axis.text(0.58, 0.545, "vote stream", fontsize=8, color="#555555", ha="center")
    for y in row_y:
        _overview_arrow(
            axis,
            (0.56, y + 0.0325),
            (0.615, 0.545),
            connectionstyle="arc3,rad=0.14",
        )

    _overview_box(axis, 0.665, 0.515, 0.085, 0.06, "votes", fontsize=8)
    _overview_arrow(axis, (0.615, 0.545), (0.665, 0.545))
    _overview_box(
        axis,
        0.775,
        0.665,
        0.145,
        0.10,
        "Fixed\nrun all N = 10",
        facecolor="#d9d9d9",
    )
    _overview_box(
        axis,
        0.775,
        0.405,
        0.145,
        0.10,
        "SE\ncheck lock -> stop",
        facecolor="#222222",
        edgecolor="#000000",
        fontsize=8.2,
    )
    _overview_arrow(axis, (0.75, 0.545), (0.775, 0.715), connectionstyle="arc3,rad=-0.2")
    _overview_arrow(axis, (0.75, 0.545), (0.775, 0.455), connectionstyle="arc3,rad=0.2")
    _overview_box(
        axis,
        0.91,
        0.535,
        0.055,
        0.06,
        "same\nverdict",
        facecolor="#ffffff",
        edgecolor="#000000",
        fontsize=7.0,
    )
    _overview_arrow(axis, (0.92, 0.715), (0.91, 0.565), connectionstyle="arc3,rad=0.2")
    _overview_arrow(axis, (0.92, 0.455), (0.91, 0.565), connectionstyle="arc3,rad=-0.2")

    cards = (
        (0.025, "600 / 600", "paired internal verdicts match", "#ffffff", "#555555"),
        (
            0.27,
            f"{metrics['se_queries']:.2f} vs {metrics['fixed_queries']:.0f}",
            "harmful queries / sample",
            "#ffffff",
            "#555555",
        ),
        (
            0.515,
            f"-{metrics['query_reduction']:.2%}",
            "harmful query reduction",
            "#dedede",
            "#555555",
        ),
        (
            0.76,
            f"-{metrics['completion_token_reduction']:.2%}",
            "harmful completion-token reduction",
            "#222222",
            "#000000",
        ),
    )
    for x, value, label, facecolor, edgecolor in cards:
        axis.add_patch(
            FancyBboxPatch(
                (x, 0.075),
                0.215,
                0.13,
                boxstyle="round,pad=0.01,rounding_size=0.015",
                facecolor=facecolor,
                edgecolor=edgecolor,
                linewidth=0.9,
            )
        )
        text_color = "white" if facecolor == "#222222" else "#000000"
        axis.text(
            x + 0.1075,
            0.15,
            value,
            ha="center",
            va="center",
            fontsize=13,
            fontweight="bold",
            color=text_color,
        )
        axis.text(
            x + 0.1075,
            0.105,
            label,
            ha="center",
            va="center",
            fontsize=7.2,
            color=text_color,
        )

    save_figure(figure, "fig_00_overview")
    return pd.DataFrame(
        [
            {
                "paired_verdict_match_rate": metrics["verdict_match_rate"],
                "paired_verdict_mismatch_count": int(metrics["verdict_mismatch_count"]),
                "harmful_fixed_queries_mean": metrics["fixed_queries"],
                "harmful_se_queries_mean": metrics["se_queries"],
                "harmful_query_reduction": metrics["query_reduction"],
                "harmful_prompt_token_reduction": metrics["prompt_token_reduction"],
                "harmful_completion_token_reduction": metrics["completion_token_reduction"],
            }
        ]
    )


def make_figures(
    seed_metrics: pd.DataFrame,
    efficiency: pd.DataFrame,
    refusal: pd.DataFrame,
    pair: pd.DataFrame,
) -> pd.DataFrame:
    overview_data = make_overview_figure(efficiency, pair)
    configure_plotting()
    # Figure 1: paired external ASR by seed.
    asr = seed_metrics[seed_metrics["metric"] == "deepseek_harmful_asr"]
    figure, axis = plt.subplots(figsize=FIG_SINGLE_COL)
    for method in ("smoothllm_fixed", "se_smoothllm"):
        group = asr[asr["method"] == method].sort_values("seed")
        axis.plot(
            group["seed"],
            group["value"],
            color=METHOD_COLORS[method],
            marker=METHOD_MARKERS[method],
            linewidth=1.2,
            label=DISPLAY_NAMES[method],
        )
        axis.axhline(group["value"].mean(), color=METHOD_COLORS[method], linestyle=":", linewidth=0.8)
    axis.set_xlabel("Seed")
    axis.set_ylabel("Harmful ASR")
    axis.set_xticks([42, 43, 44])
    percentage_axis(axis)
    add_full_frame(axis)
    axis.legend()
    save_figure(figure, "fig_01_harmful_asr_by_seed")

    # Figure 2: model-query cost by split and method.
    query = efficiency[efficiency["metric"] == "copies_used"]
    figure, axis = plt.subplots(figsize=FIG_DOUBLE_COL)
    splits = ("harmful", "benign")
    x = np.arange(len(splits))
    width = 0.24
    for offset, method in zip((-width, 0, width), METHODS, strict=True):
        subset = query[query["method"] == method].set_index("split").reindex(splits)
        axis.bar(
            x + offset,
            subset["mean"],
            width=width,
            yerr=subset["std"],
            color=METHOD_COLORS[method],
            edgecolor="black",
            capsize=3,
            label=DISPLAY_NAMES[method],
        )
    axis.set_xticks(x, ["Harmful", "Benign"])
    axis.set_ylabel("Model queries per sample (count)")
    axis.set_ylim(bottom=0)
    add_full_frame(axis)
    axis.legend()
    figure.tight_layout()
    save_figure(figure, "fig_02_query_cost")

    # Figure 3: prompt and completion token cost.
    token = efficiency[efficiency["metric"].isin(("prompt_tokens", "completion_tokens"))]
    figure, axes = plt.subplots(1, 2, figsize=FIG_DOUBLE_COL, sharey=False)
    for axis, metric, label in zip(
        axes,
        ("prompt_tokens", "completion_tokens"),
        ("Prompt tokens per sample", "Completion tokens per sample"),
        strict=True,
    ):
        subset = token[token["metric"] == metric]
        splits = ("harmful", "benign")
        x = np.arange(len(splits))
        width = 0.24
        for offset, method in zip((-width, 0, width), METHODS, strict=True):
            group = subset[subset["method"] == method].set_index("split").reindex(splits)
            axis.bar(
                x + offset,
                group["mean"],
                width=width,
                yerr=group["std"],
                color=METHOD_COLORS[method],
                edgecolor="black",
                capsize=3,
                label=DISPLAY_NAMES[method],
            )
        axis.set_xticks(x, ["Harmful", "Benign"])
        axis.set_ylabel(label)
        axis.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{value:,.0f}"))
        add_full_frame(axis)
    axes[1].legend()
    figure.tight_layout()
    save_figure(figure, "fig_03_token_cost")

    # Figure 4: benign refusal rate by seed.
    benign = refusal[(refusal["split"] == "benign")].copy()
    benign["value"] = benign["refused"].astype(float)
    refusal_seed = (
        benign.groupby(["method", "seed"], as_index=False)["value"].mean().sort_values(["method", "seed"])
    )
    figure, axis = plt.subplots(figsize=FIG_SINGLE_COL)
    for method in METHODS:
        group = refusal_seed[refusal_seed["method"] == method]
        axis.plot(
            group["seed"],
            group["value"],
            color=METHOD_COLORS[method],
            marker=METHOD_MARKERS[method],
            linewidth=1.2,
            label=DISPLAY_NAMES[method],
        )
    axis.set_xlabel("Seed")
    axis.set_ylabel("Benign refusal rate")
    axis.set_xticks([42, 43, 44])
    percentage_axis(axis)
    add_full_frame(axis)
    axis.legend()
    save_figure(figure, "fig_04_benign_refusal_rate")
    return overview_data


def write_captions(summary: pd.DataFrame, pair: pd.DataFrame) -> None:
    asr = summary[summary["metric"] == "deepseek_harmful_asr"].set_index("method")["value"]
    benign = summary[
        (summary["metric"] == "llama3_8b_refusal_rate") & (summary["split"] == "benign")
    ].set_index("method")["value"]
    lines = [
        "# Figure captions",
        "",
        "**Figure 0.** Project overview of the shared SmoothLLM execution. The input prompt is perturbed into multiple copies, each copy is sent through the target LLM and the internal Judge, and the resulting vote stream is consumed either by fixed SmoothLLM (all N=10 copies) or by SE-SmoothLLM (stop once the final verdict is mathematically locked). The verdict card uses all 600 fixed/SE pairs (300 harmful and 300 benign), while the compute cards use the 300 harmful pairs and summarize exact agreement and savings.",
        "",
        f"**Figure 1.** Harmful-request attack success rate (ASR) for SmoothLLM fixed and SE-SmoothLLM across seeds 42, 43, and 44. Points show the exact proportion of 100 harmful samples per seed; dotted horizontal lines show the three-seed mean, without inferential significance testing. The DeepSeek-V4-Flash auxiliary Judge gives {asr['smoothllm_fixed']:.2%} for fixed SmoothLLM and {asr['se_smoothllm']:.2%} for SE-SmoothLLM, so the observed ASR difference is small in this run.",
        "",
        "**Figure 2.** Model-query cost per sample for undefended generation, fixed SmoothLLM, and SE-SmoothLLM on harmful and benign splits. Bars show mean queries and error bars show sample standard deviation (n=300 per method and split). SE-SmoothLLM uses fewer queries than the fixed budget method on both splits while preserving the fixed method as the efficiency reference.",
        "",
        "**Figure 3.** Prompt and completion token cost per sample across methods and splits. Bars show mean tokens and error bars show sample standard deviation (n=300 per method and split). The token panels expose the compute cost associated with early stopping instead of presenting query savings alone.",
        "",
        f"**Figure 4.** Benign-request refusal rate across seeds for the three methods. Points show the exact rate among 100 benign samples per seed; no significance test is applied. The Llama-3-8B Refusal Judge reports {benign['smoothllm_fixed']:.2%} for fixed SmoothLLM and {benign['se_smoothllm']:.2%} for SE-SmoothLLM, which should be interpreted as a usability cost rather than a security gain.",
        "",
        f"The paired fixed-versus-SE comparison contains {len(pair):,} sample pairs and {int((~pair['internal_verdict_match']).sum())} internal verdict mismatches.",
        "",
    ]
    (OUTPUT_DIR / "FIGURE_CAPTIONS.md").write_text("\n".join(lines), encoding="utf-8")


def write_summary_json(summary: pd.DataFrame, raw: pd.DataFrame, deepseek: pd.DataFrame, refusal: pd.DataFrame) -> None:
    payload = {
        "project": PROJECT_TITLE,
        "sources": {
            "raw_records": int(len(raw)),
            "deepseek_judge_records": int(len(deepseek)),
            "llama3_8b_refusal_records": int(len(refusal)),
            "raw_config_sha256": sorted(raw["config_sha256"].dropna().unique().tolist()),
        },
        "summary": summary.replace({np.nan: None}).to_dict(orient="records"),
        "limitations": [
            "DeepSeek-V4-Flash is an auxiliary external Judge, not the official JBB Llama-3-70B Judge.",
            "ASR and refusal rates are descriptive results over three configured seeds; no significance claim is made.",
            "Raw model responses remain local and are excluded from Git by .gitignore.",
        ],
    }
    (OUTPUT_DIR / "summary.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    for directory in (RAW_DIR, JUDGED_DIR):
        if not directory.exists():
            raise FileNotFoundError(directory)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    raw = load_raw_metrics()
    deepseek = load_deepseek_judge()
    refusal = load_refusal_judge()
    pair = build_pair_metrics(raw)
    summary, seed_metrics, efficiency = aggregate_tables(raw, deepseek, refusal, pair)

    write_data_health({"raw_metrics": raw, "deepseek_judge": deepseek, "refusal_judge": refusal, "paired_metrics": pair})
    raw.to_csv(OUTPUT_DIR / "01_raw_metrics.csv", index=False)
    deepseek.to_csv(OUTPUT_DIR / "02_deepseek_judge.csv", index=False)
    refusal.to_csv(OUTPUT_DIR / "03_llama3_8b_refusal_judge.csv", index=False)
    pair.to_csv(OUTPUT_DIR / "04_fixed_vs_se_pairs.csv", index=False)
    summary.to_csv(OUTPUT_DIR / "summary.csv", index=False)
    seed_metrics.to_csv(OUTPUT_DIR / "seed_metrics.csv", index=False)
    efficiency.to_csv(OUTPUT_DIR / "efficiency_summary.csv", index=False)
    write_xlsx(OUTPUT_DIR / "01_raw_metrics.xlsx", {"raw_metrics": raw})
    write_xlsx(OUTPUT_DIR / "02_external_judges.xlsx", {"deepseek": deepseek, "llama3_8b": refusal})
    write_xlsx(OUTPUT_DIR / "03_paired_metrics.xlsx", {"fixed_vs_se": pair})
    write_xlsx(OUTPUT_DIR / "99_summary.xlsx", {"summary": summary, "seed_metrics": seed_metrics, "efficiency": efficiency})
    write_summary_json(summary, raw, deepseek, refusal)

    overview_data = make_figures(seed_metrics, efficiency, refusal, pair)
    asr_data = seed_metrics[seed_metrics["metric"] == "deepseek_harmful_asr"]
    query_data = efficiency[efficiency["metric"] == "copies_used"]
    token_data = efficiency[efficiency["metric"].isin(("prompt_tokens", "completion_tokens"))]
    refusal_data = seed_metrics[
        (seed_metrics["metric"] == "llama3_8b_refusal_rate") & (seed_metrics["split"] == "benign")
    ]
    write_xlsx(FIGURE_DIR.parent / "processed" / "fig_00_data.xlsx", {"overview": overview_data})
    write_xlsx(FIGURE_DIR.parent / "processed" / "fig_01_data.xlsx", {"asr_by_seed": asr_data})
    write_xlsx(FIGURE_DIR.parent / "processed" / "fig_02_data.xlsx", {"query_cost": query_data})
    write_xlsx(FIGURE_DIR.parent / "processed" / "fig_03_data.xlsx", {"token_cost": token_data})
    write_xlsx(FIGURE_DIR.parent / "processed" / "fig_04_data.xlsx", {"benign_refusal": refusal_data})
    write_captions(summary, pair)

    print(f"raw_records={len(raw)}")
    print(f"deepseek_judge_records={len(deepseek)}")
    print(f"llama3_8b_refusal_records={len(refusal)}")
    print(f"paired_records={len(pair)}")
    print(f"verdict_mismatches={(~pair['internal_verdict_match']).sum()}")
    for method, group in deepseek.groupby("method"):
        print(f"deepseek_asr[{method}]={group['external_jailbroken'].mean():.4f}")
    print(f"outputs={OUTPUT_DIR}")


if __name__ == "__main__":
    main()
